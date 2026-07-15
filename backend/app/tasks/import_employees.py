import json
import io

from openpyxl import load_workbook

from app.core.database import SessionLocal
from app.core.redis import redis_client
from app.core.celery_app import celery_app
from app.repositories import employee_repo, operation_log_repo

_FIELDS = ["employee_no", "name", "gender", "department_id", "position",
           "direct_supervisor_id", "phone", "email", "entry_date", "status"]


@celery_app.task(name="app.tasks.import_employees.run")
def run_import(file_data: bytes, user_id: str, task_id: str) -> None:
    wb = load_workbook(io.BytesIO(file_data), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        _save(task_id, "failed", None, "Excel 文件为空")
        return
    headers = [str(h).strip() if h else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}
    errors: list[dict] = []
    success = 0
    db = SessionLocal()
    try:
        for i, row in enumerate(rows[1:], start=2):
            vals = {}
            for f in _FIELDS:
                if f in idx and idx[f] < len(row):
                    vals[f] = row[idx[f]]
            err = _validate_row(vals, i, db)
            if err:
                errors.append(err)
                continue
            _create_employee(vals, db)
            success += 1
        operation_log_repo.create(db, user_id=user_id, module="employee", action="import",
                                  target_type="employee", target_id=None,
                                  detail={"success": success, "fail": len(errors), "errors": errors})
        db.commit()
        _save(task_id, "success", {"success_count": success, "error_count": len(errors), "errors": errors}, None)
    except Exception as e:
        db.rollback()
        _save(task_id, "failed", None, str(e))
    finally:
        db.close()


def _validate_row(vals: dict, row_num: int, db) -> dict | None:
    if not vals.get("employee_no") or not vals.get("name"):
        return {"row": row_num, "error": "工号和姓名为必填"}
    eno = str(vals["employee_no"]).strip()
    if employee_repo.get_by_employee_no(db, eno) is not None:
        return {"row": row_num, "error": f"工号 {eno} 已存在"}
    try:
        gender = int(vals.get("gender", 1))
        if gender not in (0, 1, 2):
            return {"row": row_num, "error": "性别必须为 0/1/2"}
    except (ValueError, TypeError):
        return {"row": row_num, "error": "性别格式错误"}
    dept_id = vals.get("department_id")
    if dept_id:
        from app.repositories import department_repo
        if department_repo.get_active(db, str(dept_id)) is None:
            return {"row": row_num, "error": f"部门 {dept_id} 不存在"}
    sup_id = vals.get("direct_supervisor_id")
    if sup_id:
        sup = employee_repo.get_active(db, str(sup_id))
        if sup is None:
            return {"row": row_num, "error": f"上级 {sup_id} 不存在"}
    return None


def _create_employee(vals: dict, db) -> None:
    fields = {"employee_no": str(vals["employee_no"]).strip(),
              "name": str(vals["name"]).strip(),
              "gender": int(vals.get("gender", 1)),
              "status": int(vals.get("status", 1))}
    for f in ["department_id", "position", "direct_supervisor_id", "phone", "email"]:
        v = vals.get(f)
        if v is not None and str(v).strip() != "":
            fields[f] = str(v).strip()
    entry = vals.get("entry_date")
    if entry:
        try:
            from datetime import datetime
            if isinstance(entry, datetime):
                fields["hire_date"] = entry.date()
            elif isinstance(entry, str):
                fields["hire_date"] = datetime.strptime(entry, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            pass
    employee_repo.create(db, **fields)


def _save(task_id: str, status: str, result, error: str | None) -> None:
    redis_client.set(f"task:{task_id}",
                     json.dumps({"status": status, "result": result, "error": error}, ensure_ascii=False),
                     ex=3600)
